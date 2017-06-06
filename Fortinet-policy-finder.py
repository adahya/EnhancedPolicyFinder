import os
import sys
import copy
import argparse
import re
import pprint
import ipaddress
import json
import openpyxl
from enum import Enum

targetsubnetsdict = dict()
addrobjdict = dict()
addrgrpobjdict = dict()
interfaceobjdict = dict()
ServiceCutobjdict = dict()
ServiceGrpobjdict = dict()
Policyobject = dict()
StaticRoutingobject = dict()
MatchingConfigDict = dict()

class AddrObjT(Enum):
	Subnet = 1
	IPRange = 2
	unknown = 127

#Target Is in this Route
def Is_inRoute(target,Routerobj):
	IPTarget = ipaddress.ip_network(str(target))
	IPRouteObj = ipaddress.ip_network(str(Routerobj['dst']).replace(' ','/'))
	#print("Original : " + str(Routerobj['dst']) + " Now is : " + str(IPRouteObj))
	if IPTarget.overlaps(IPRouteObj) is True:
			return True
	return False
	
#Target Is in this Subnets
def Is_inSubnet(target,addrobj):
	IPTarget = ipaddress.ip_network(str(target))
	IPsubnetObj = ipaddress.ip_network(str(addrobj['subnet']).replace(' ','/'))
	#print("Original : " + str(addrobj['subnet']) + " Now is : " + str(IPsubnetObj))
	if IPTarget.overlaps(IPsubnetObj) is True:
			return True
	return False

#Target Is in this range
def Is_inRange(target,addrobj):
	IPTarget = ipaddress.ip_network(str(target))
	IPStartObj = ipaddress.IPv4Address(str(addrobj['start-ip']))
	IPEndObj = ipaddress.IPv4Address(str(addrobj['end-ip']))
	temp = IPStartObj
	#print ( "Start with " + str(IPStartObj) + " End with " + str(IPEndObj)) 
	while not temp > IPEndObj:
		if temp in IPTarget:
			return True
		temp+= 1
	return False
	
#Determine the type of Address Objects
def AddrObjType(addrobj):
	if 'subnet' in addrobj:
		return (AddrObjT.Subnet)
	elif 'type' in addrobj and addrobj['type'] in 'iprange':
		return (AddrObjT.IPRange)
	else :
		return (AddrObjT.unknown)

#Match MatchAddrObject(targetsubnetsdict)
def MatchAddrObject(targetsubnets):
	MatchingConfigDict['AddressObject'] = dict()
	for target in targetsubnets:
		#print (targetsubnets[target])
		for addrObj in addrobjdict:
			AddrType = AddrObjType(addrobjdict[addrObj])
			if AddrType is AddrObjT.Subnet and Is_inSubnet(targetsubnets[target],addrobjdict[addrObj]) is True:
				#print ("Match the Object " + str(addrobjdict[addrObj]))
				MatchingConfigDict['AddressObject'][addrObj] = addrobjdict[addrObj].copy()
			elif AddrType is AddrObjT.IPRange and Is_inRange(targetsubnets[target],addrobjdict[addrObj]) is True:
				MatchingConfigDict['AddressObject'][addrObj] = addrobjdict[addrObj].copy()

#Match MatchAddrGrpObject(targetsubnetsdict) 
def MatchAddrGrpObject(targetsubnets):
	MatchingConfigDict['AddressGrpObject'] = dict()
	for key in targetsubnets.keys():
		#print (target)
		for addrGrp in addrgrpobjdict:
			if key in addrgrpobjdict[addrGrp]['member']:
				#print(addrgrpobjdict[addrGrp])
				MatchingConfigDict['AddressGrpObject'][addrGrp] = addrgrpobjdict[addrGrp].copy()

#Match MatchPoliciesObject(targetsubnetsdict) Policyobject
def MatchPoliciesObject(targetsubnets):
	MatchingConfigDict['PolicyObject'] = dict()
	for key in targetsubnets['AddressObject'].keys():
		#print (key)
		for Policyid in Policyobject:
			if key in Policyobject[Policyid]['srcaddr'] or key in Policyobject[Policyid]['dstaddr']:
				#print(Policyobject[Policyid])
				MatchingConfigDict['PolicyObject'][Policyid] = Policyobject[Policyid].copy()
				
#Match MatchRouteObject(targetsubnetsdict)
def MatchRouteObject(targetsubnets):
	MatchingConfigDict['RouteObject'] = dict()
	for target in targetsubnets:
		#print (targetsubnets[target])
		for RouteObj in StaticRoutingobject:
			#print(StaticRoutingobject[RouteObj])
			if Is_inRoute(targetsubnets[target],StaticRoutingobject[RouteObj]) is True:
				MatchingConfigDict['RouteObject'][RouteObj] = StaticRoutingobject[RouteObj].copy()

#Match MatchCustomSrvObject(MatchingConfigDict['PolicyObject'])
def MatchCustomSrvObject(targetsubnets):
	MatchingConfigDict['CustomSrvObject'] = dict()
	for service in ServiceCutobjdict:
		#print(service)
		for SrviceKey in MatchingConfigDict['PolicyObject']:
			#print(MatchingConfigDict['PolicyObject'][SrviceKey])
			if service in MatchingConfigDict['PolicyObject'][SrviceKey]['service']:
					MatchingConfigDict['CustomSrvObject'][service] = ServiceCutobjdict[service].copy()

#Match MatchCustomSrvGrpObject(MatchingConfigDict['PolicyObject']) 
def MatchCustomSrvGrpObject(targetsubnets):
	MatchingConfigDict['CustomSrvGrpObject'] = dict()
	for serviceGrp in ServiceGrpobjdict:
		#print(serviceGrp)
		for SrviceKey in MatchingConfigDict['PolicyObject']:
			#print(MatchingConfigDict['PolicyObject'][SrviceKey])
			if serviceGrp in MatchingConfigDict['PolicyObject'][SrviceKey]['service']:
					MatchingConfigDict['CustomSrvGrpObject'][serviceGrp] = ServiceGrpobjdict[serviceGrp].copy()
					
#process address objected			
def ProcessAddressObject(fullconfiglines):		
	allobject = fullconfiglines[fullconfiglines.index('config firewall address') + 1:]
	allobject = allobject[:allobject.index('end')]
	
	for line in allobject:
		try:
			if line.strip().startswith('edit'):
				objid = (re.match(r'edit \"(.*)\"', line.strip()).groups()[0])
				addrobjdict[objid] = dict()
			elif line.strip() != 'next' and line.strip().startswith('set'):
				key, val = re.match(r'^set (\S*) (.+)$', line.strip()).groups()
				addrobjdict[objid][key] = val #.replace('" "','')
		except:
			print(("Error on line: %s" % line))
			raise
	if 'all' in addrobjdict and 'subnet' not in addrobjdict['all']:
		addrobjdict['all']["subnet"] = "0.0.0.0 0.0.0.0"
		print (addrobjdict['all'])
	
#process Interfaces objected			
def ProcessInterfaceObject(fullconfiglines):		
	allobject = fullconfiglines[fullconfiglines.index('config system interface') + 1:]
	allobject = allobject[:allobject.index('end')]
	
	for line in allobject:
		try:
			if line.strip().startswith('edit'):
				objid = (re.match(r'edit \"(.*)"', line.strip()).groups()[0])
				interfaceobjdict[objid] = dict()
			elif line.strip() != 'next' and line.strip().startswith('set'):
				key, val = re.match(r'^set (\S*) (.+)$', line.strip()).groups()
				interfaceobjdict[objid][key] = val #.replace('" "','')
		except:
			print(("Error on line: %s" % line))
			raise

#process Address Group objected			
def ProcessaddrGrpObject(fullconfiglines):		
	allobject = fullconfiglines[fullconfiglines.index('config firewall addrgrp') + 1:]
	allobject = allobject[:allobject.index('end')]
	
	for line in allobject:
		try:
			if line.strip().startswith('edit'):
				objid = (re.match(r'edit \"(.*)"', line.strip()).groups()[0])
				addrgrpobjdict[objid] = dict()
			elif line.strip() != 'next' and line.strip().startswith('set'):
				key, val = re.match(r'^set (\S*) (.+)$', line.strip()).groups()
				addrgrpobjdict[objid][key] = val #.replace('" "','')
		except:
			print(("Error on line: %s" % line))
			raise
	
#process Custom Services objected			
def ProcessSrvCutGrpObject(fullconfiglines):		
 
	allobject = fullconfiglines[fullconfiglines.index('config firewall service custom') + 1:]
	allobject = allobject[:allobject.index('end')]
	
	for line in allobject:
		try:
			if line.strip().startswith('edit'):
				objid = (re.match(r'edit \"(.*)"', line.strip()).groups()[0])
				ServiceCutobjdict[objid] = dict()
			elif line.strip() != 'next' and line.strip().startswith('set'):
				key, val = re.match(r'^set (\S*) (.+)$', line.strip()).groups()
				ServiceCutobjdict[objid][key] = val #.replace('" "','')
		except:
			print(("Error on line: %s" % line))
			raise
			
#process Custom Services objected			
def ProcessSrvGrpGrpObject(fullconfiglines):		
		
	allobject = fullconfiglines[fullconfiglines.index('config firewall service group') + 1:]
	allobject = allobject[:allobject.index('end')]
	
	for line in allobject:
		try:
			if line.strip().startswith('edit'):
				objid = (re.match(r'edit \"(.*)\"', line.strip()).groups()[0])
				ServiceGrpobjdict[objid] = dict()
			elif line.strip() != 'next' and line.strip().startswith('set'):
				key, val = re.match(r'^set (\S*) (.+)$', line.strip()).groups()
				ServiceGrpobjdict[objid][key] = val #.replace('" "','')
		except:
			print(("Error on line: %s" % line))
			raise
			
#process Policies objected			
def ProcessPoliciesGrpObject(fullconfiglines):	

	allobject = fullconfiglines[fullconfiglines.index('config firewall policy') + 1:]
	allobject = allobject[:allobject.index('end')]
	
	for line in allobject:
		try:
			if line.strip().startswith('edit'):
				objid = (re.match(r'edit (.*)', line.strip()).groups()[0])
				Policyobject[objid] = dict()
			elif line.strip() != 'next' and line.strip().startswith('set'):
				key, val = re.match(r'^set (\S*) (.+)$', line.strip()).groups()
				Policyobject[objid][key] = val #.replace('" "','')
		except:
			print(("Error on line: %s" % line))
			raise
			
#process Static Routing objected			
def ProcessStaticRoutingObject(fullconfiglines):	

	allobject = fullconfiglines[fullconfiglines.index('config router static') + 1:]
	allobject = allobject[:allobject.index('end')]
	
	for line in allobject:
		try:
			if line.strip().startswith('edit'):
				objid = (re.match(r'edit (.*)', line.strip()).groups()[0])
				StaticRoutingobject[objid] = dict()
			elif line.strip() != 'next' and line.strip().startswith('set'):
				key, val = re.match(r'^set (\S*) (.+)$', line.strip()).groups()
				StaticRoutingobject[objid][key] = val #.replace('" "','')
		except:
			print(("Error on line: %s" % line))
			raise
	for routes in StaticRoutingobject:
		if 'dst' not in StaticRoutingobject[routes]:
			StaticRoutingobject[routes]["dst"] = "0.0.0.0 0.0.0.0"
		
		
# Print Dict of file name
def printfilenames(filenamelist):
	for id in filenamelist:
		print("     [" + str(id) + "] : " + filenamelist[id])


# find configuration file names
def findconfiglist(option=None):
	#define local variables
	filelist_dict = dict()
	files = os.listdir(".")
	i = 0
	for file in files:
		m = re.match(r'^(.*)\.conf$', file)
		if m != None:
			i += 1
			filelist_dict[i] = file
	return filelist_dict

	
	
if __name__ == "__main__":

	print(("Program Running .... \n"))
	filelist = findconfiglist()
	print("Found configuration file :")
	printfilenames(filelist)
	configfileid = input("Please enter id of source config file: ")
	print("Selected ID : " + configfileid + "  Filename : " + filelist[int(configfileid)])
	
	#Creating Dictionary for Address Objects
	print("PHASE 1: Loading Configuration File " + filelist[int(configfileid)])
	try:
		fullconfigstr = open(filelist[int(configfileid)], 'r', encoding='UTF8').read()
	except FileNotFoundError:
		print("Error reading config file: %s" + filelist[int(configfileid)])
		sys.exit(FileNotFoundError.errno)
	fullconfiglines = fullconfigstr.splitlines()
	
	ProcessInterfaceObject(fullconfiglines)
	ProcessStaticRoutingObject(fullconfiglines)
	ProcessAddressObject(fullconfiglines)
	ProcessaddrGrpObject(fullconfiglines)
	ProcessSrvCutGrpObject(fullconfiglines)
	ProcessSrvGrpGrpObject(fullconfiglines)
	ProcessPoliciesGrpObject(fullconfiglines)
		
	fd = open('test-output.json', 'w')
	fd.write(json.dumps({
		"addrobjdict": addrobjdict,
		"addrgrpobjdict": addrgrpobjdict,
		"interfaceobjdict": interfaceobjdict,
		"ServiceCutobjdict": ServiceCutobjdict,
		"ServiceGrpobjdict": ServiceGrpobjdict,
		"Policyobject": Policyobject,
		"StaticRoutingobject": StaticRoutingobject,
		}, sort_keys=True,indent=4,separators=(',', ': ')))
	fd.close()
	
	filename = ""
	while filename == "":
		filename = 'full.xlsx'
		print("Checking Subnets in File full.xlsx)")
	print ("PHASE 2: Openging Excel Workbook " + filename + " and Loading Targeted subnets")
	wb = openpyxl.load_workbook('full.xlsx')
	sheet = wb.get_sheet_by_name('Matching Subnets')
	
	#targetsubnetsdict
	row = 1
	matched=0
	while sheet.cell(row=row, column=1).value!=None:
		targetsubnetsdict[row]= sheet.cell(row=row, column=1).value
		row+= 1
		
	MatchAddrObject(targetsubnetsdict)
	MatchAddrGrpObject(MatchingConfigDict['AddressObject'])
	MatchPoliciesObject(MatchingConfigDict)
	MatchRouteObject(targetsubnetsdict)
	MatchCustomSrvObject(MatchingConfigDict['PolicyObject'])
	MatchCustomSrvGrpObject(MatchingConfigDict['PolicyObject'])
	MatchingConfigDict['Statistics'] = dict()
	MatchingConfigDict['Statistics']['Address Objects'] = len(MatchingConfigDict['AddressObject'])
	MatchingConfigDict['Statistics']['Address Groups'] = len(MatchingConfigDict['AddressGrpObject'])
	MatchingConfigDict['Statistics']['Firewall Policies'] = len(MatchingConfigDict['PolicyObject'])
	MatchingConfigDict['Statistics']['Service Objects'] = len(MatchingConfigDict['CustomSrvObject'])
	MatchingConfigDict['Statistics']['Service Groups'] = len(MatchingConfigDict['CustomSrvGrpObject'])
	MatchingConfigDict['Statistics']['Static Routes'] = len(MatchingConfigDict['RouteObject'])
	#print(MatchingConfigDict)
	
	fd = open('Matches-output-' + filelist[int(configfileid)] + '.json', 'w')
	fd.write(json.dumps({
		"Statistics": MatchingConfigDict['Statistics'],
		"config firewall address": MatchingConfigDict['AddressObject'],
		"config firewall addrgrp": MatchingConfigDict['AddressGrpObject'],
		"config firewall policy": MatchingConfigDict['PolicyObject'],
		"config firewall service custom": MatchingConfigDict['CustomSrvObject'],
		"config firewall service group": MatchingConfigDict['CustomSrvGrpObject'],
		"config router static": MatchingConfigDict['RouteObject'],
		}, sort_keys=True,indent=4,separators=(',', ': ')))
	fd.close()
	
	

